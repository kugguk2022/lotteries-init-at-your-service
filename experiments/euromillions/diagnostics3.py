from __future__ import annotations

import argparse
import json
import math
from dataclasses import asdict, dataclass
from itertools import combinations, islice
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
from scipy.stats import norm

from euromillions.schema import validate_df
from euromillions_agent.phase2_sobol import euler_phi_upto

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover - environment fallback
    Workbook = None

try:
    import xlsxwriter
except Exception:  # pragma: no cover - environment fallback
    xlsxwriter = None

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_HISTORY = REPO_ROOT / "data" / "euromillions.csv"
DEFAULT_OUT_DIR = REPO_ROOT / "outputs" / "euromillions" / "diagnostics3"
DEFAULT_BATCH_SIZE = 200_000
DEFAULT_MODE = "full7"
DEFAULT_ROLLING_WINDOW = 100
EXCEL_MAX_ROWS = 1_048_576
FULL7_DEFAULT_START_DATE = "2016-09-27"


@dataclass
class Diagnostics3Summary:
    raw_rows: int
    mode: str
    rows: int
    cutoff_start_date: str
    start_date: str
    end_date: str
    main_n: int
    star_n: int
    pair_dimension: int
    phi_poi_corr: float
    poi_mean: float
    poi_std: float
    phi_model: str
    glm_family: str
    glm_link: str
    model_intercept: float
    model_phi_linear: float
    model_phi_quadratic: float
    model_aic: float
    model_deviance: float
    model_fitted_mean: float
    model_shifted_prediction_mean: float
    pruning_applied: bool
    pruning_z_threshold: float
    inlier_count: int
    outlier_count: int
    current_expected_poi: float
    predicted_next_poi: float
    predicted_growth_n1: float
    reverse_growth_target_score: int
    reverse_growth_target_value: float
    reverse_growth_effective_score: int
    reverse_growth_score_gap: int
    trailing_window_r_compatible: int
    trailing_mean_r_compatible: float
    trailing_mean_last_5: float
    target_score_r_compatible: int
    matching_guess_count: int
    total_combinations_scored: int
    reverse_growth_match_count: int
    least_likely_main_pair: list[int]
    least_likely_main_pair_z: float
    most_likely_main_pair: list[int]
    most_likely_main_pair_z: float


def resolve_repo_path(path: Path) -> Path:
    return path if path.is_absolute() else REPO_ROOT / path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Diagnostics 3: Euler-phi GLM and guess extraction for EuroMillions."
    )
    parser.add_argument(
        "--history",
        type=Path,
        default=DEFAULT_HISTORY,
        help="Normalized EuroMillions history CSV with draw_date and ball_1..ball_5.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=DEFAULT_OUT_DIR,
        help="Directory for diagnostics 3 CSV, JSON, and plot outputs.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help="Number of 5-ball combinations to score per batch.",
    )
    parser.add_argument(
        "--mode",
        choices=("full7", "main5"),
        default=DEFAULT_MODE,
        help="Use the full 5+2 ticket or the original 5-main-only R shape.",
    )
    parser.add_argument(
        "--outlier-z",
        type=float,
        default=3.5,
        help="Robust z-threshold for Euler-totient residual pruning.",
    )
    parser.add_argument(
        "--rolling-window",
        type=int,
        default=DEFAULT_ROLLING_WINDOW,
        help="Rolling window for order-statistic diagnostics.",
    )
    parser.add_argument(
        "--excel-top-n",
        type=int,
        default=500,
        help="Number of reverse-growth candidates to keep in the Excel shortlist sheet.",
    )
    parser.add_argument(
        "--start-date",
        type=str,
        default=None,
        help=(
            "Optional inclusive draw-date cutoff in YYYY-MM-DD. "
            "If omitted, full7 mode defaults to 2016-09-27 so phi predictions stay in the 12-star era."
        ),
    )
    return parser.parse_args()


def load_history(history_path: Path) -> pd.DataFrame:
    history = pd.read_csv(history_path)
    history = validate_df(history)
    history = history.sort_values("draw_date").drop_duplicates(subset=["draw_date"]).reset_index(drop=True)

    required = [f"ball_{idx}" for idx in range(1, 6)] + [f"star_{idx}" for idx in range(1, 3)]
    missing = [col for col in required if col not in history.columns]
    if missing:
        raise ValueError(
            f"History file is missing expected columns: {missing}. "
            "Expected normalized columns ball_1..ball_5 and star_1..star_2."
        )
    return history


def apply_start_date_cutoff(
    history: pd.DataFrame, *, mode: str, start_date_arg: str | None
) -> tuple[pd.DataFrame, str]:
    effective_start = start_date_arg
    if effective_start is None and mode == "full7":
        effective_start = FULL7_DEFAULT_START_DATE
    if effective_start is None:
        return history, history["draw_date"].min().date().isoformat()

    cutoff = pd.Timestamp(effective_start)
    filtered = history[history["draw_date"] >= cutoff].reset_index(drop=True)
    if filtered.empty:
        raise ValueError(f"No rows remain after applying start-date cutoff {effective_start}.")
    return filtered, cutoff.date().isoformat()


@dataclass
class PairFeatureOut:
    output_pairs: np.ndarray
    poi: np.ndarray
    pair_counts: np.ndarray


@dataclass
class PairZDiagnostics:
    main_main_counts: np.ndarray
    main_main_z: np.ndarray
    main_star_counts: np.ndarray
    main_star_z: np.ndarray
    star_star_counts: np.ndarray
    star_star_z: np.ndarray
    least_main_pair: tuple[int, int]
    least_main_pair_z: float
    most_main_pair: tuple[int, int]
    most_main_pair_z: float
    top_abs_pairs: pd.DataFrame


@dataclass
class OrderStatMoments:
    means: np.ndarray
    stds: np.ndarray


def safe_zscore(values: np.ndarray) -> np.ndarray:
    arr = np.asarray(values, dtype=float)
    sigma = float(arr.std(ddof=0))
    if sigma < 1e-12:
        return np.zeros_like(arr)
    return (arr - float(arr.mean())) / sigma


def order_stat_pmf(population_size: int, sample_size: int, order_k: int) -> tuple[np.ndarray, np.ndarray]:
    denom = math.comb(population_size, sample_size)
    low = order_k
    high = population_size - sample_size + order_k
    support = np.arange(low, high + 1, dtype=int)
    probs = np.array(
        [
            math.comb(value - 1, order_k - 1)
            * math.comb(population_size - value, sample_size - order_k)
            / denom
            for value in support
        ],
        dtype=float,
    )
    return support.astype(float), probs


def main_order_stat_moments(population_size: int = 50, sample_size: int = 5) -> OrderStatMoments:
    means: list[float] = []
    stds: list[float] = []
    for order_k in range(1, sample_size + 1):
        support, probs = order_stat_pmf(population_size, sample_size, order_k)
        mean = float(np.sum(support * probs))
        var = float(np.sum(((support - mean) ** 2) * probs))
        means.append(mean)
        stds.append(var**0.5)
    return OrderStatMoments(means=np.array(means, dtype=float), stds=np.array(stds, dtype=float))


def build_pair_z_diagnostics(history: pd.DataFrame) -> PairZDiagnostics:
    mains = history[[f"ball_{idx}" for idx in range(1, 6)]].to_numpy(dtype=int)
    stars = history[[f"star_{idx}" for idx in range(1, 3)]].to_numpy(dtype=int)
    n_draws = len(history)

    main_main_counts = np.zeros((51, 51), dtype=np.int32)
    main_star_counts = np.zeros((51, 13), dtype=np.int32)
    star_star_counts = np.zeros((13, 13), dtype=np.int32)

    for draw_idx in range(n_draws):
        draw_mains = np.sort(mains[draw_idx])
        draw_stars = np.sort(stars[draw_idx])

        for left, right in combinations(draw_mains.tolist(), 2):
            main_main_counts[left, right] += 1
            main_main_counts[right, left] += 1

        for main_value in draw_mains.tolist():
            for star_value in draw_stars.tolist():
                main_star_counts[main_value, star_value] += 1

        star_left, star_right = draw_stars.tolist()
        star_star_counts[star_left, star_right] += 1
        star_star_counts[star_right, star_left] += 1

    p_main_main = math.comb(48, 3) / math.comb(50, 5)
    p_main_star = (5 / 50) * (2 / 12)
    p_star_star = 1 / math.comb(12, 2)

    def counts_to_z(counts: np.ndarray, probability: float) -> np.ndarray:
        mu = n_draws * probability
        sigma = max((n_draws * probability * (1.0 - probability)) ** 0.5, 1e-12)
        return (counts.astype(float) - mu) / sigma

    main_main_z = counts_to_z(main_main_counts, p_main_main)
    main_star_z = counts_to_z(main_star_counts, p_main_star)
    star_star_z = counts_to_z(star_star_counts, p_star_star)

    np.fill_diagonal(main_main_z, np.nan)
    np.fill_diagonal(star_star_z, np.nan)

    pair_rows: list[dict[str, float | int]] = []
    least_pair = (1, 2)
    most_pair = (1, 2)
    least_z = float("inf")
    most_z = float("-inf")
    for left in range(1, 51):
        for right in range(left + 1, 51):
            z_value = float(main_main_z[left, right])
            count_value = int(main_main_counts[left, right])
            pair_rows.append(
                {
                    "ball_1": left,
                    "ball_2": right,
                    "count": count_value,
                    "z_score": z_value,
                    "abs_z_score": abs(z_value),
                }
            )
            if z_value < least_z:
                least_z = z_value
                least_pair = (left, right)
            if z_value > most_z:
                most_z = z_value
                most_pair = (left, right)

    top_abs_pairs = (
        pd.DataFrame(pair_rows)
        .sort_values(["abs_z_score", "z_score"], ascending=[False, False])
        .reset_index(drop=True)
    )

    return PairZDiagnostics(
        main_main_counts=main_main_counts,
        main_main_z=main_main_z,
        main_star_counts=main_star_counts,
        main_star_z=main_star_z,
        star_star_counts=star_star_counts,
        star_star_z=star_star_z,
        least_main_pair=least_pair,
        least_main_pair_z=least_z,
        most_main_pair=most_pair,
        most_main_pair_z=most_z,
        top_abs_pairs=top_abs_pairs,
    )


def build_pair_features_generic(
    draws: np.ndarray,
    universe_size: int,
    *,
    include_current: bool = True,
) -> PairFeatureOut:
    draws = np.asarray(draws, dtype=int)
    n_rows, width = draws.shape
    pair_idx = list(combinations(range(width), 2))
    pair_counts = np.zeros((universe_size + 1, universe_size + 1), dtype=np.int32)
    output_pairs = np.zeros((n_rows, len(pair_idx)), dtype=np.int32)

    for row_idx in range(n_rows):
        draw = np.sort(draws[row_idx])
        if include_current:
            for a, b in combinations(draw.tolist(), 2):
                pair_counts[a, b] += 1
                pair_counts[b, a] += 1

        for pair_pos, (left_idx, right_idx) in enumerate(pair_idx):
            left = int(draw[left_idx])
            right = int(draw[right_idx])
            output_pairs[row_idx, pair_pos] = pair_counts[left, right]

        if not include_current:
            for a, b in combinations(draw.tolist(), 2):
                pair_counts[a, b] += 1
                pair_counts[b, a] += 1

    poi = output_pairs.sum(axis=1).astype(float)
    return PairFeatureOut(output_pairs=output_pairs, poi=poi, pair_counts=pair_counts)


def encode_full7_draws(history: pd.DataFrame) -> tuple[np.ndarray, int, int]:
    mains = history[[f"ball_{idx}" for idx in range(1, 6)]].to_numpy(dtype=int)
    stars = history[[f"star_{idx}" for idx in range(1, 3)]].to_numpy(dtype=int)
    main_n = int(mains.max())
    star_n = int(stars.max())
    encoded_stars = stars + main_n
    draws = np.concatenate([mains, encoded_stars], axis=1)
    return draws, main_n, star_n


def robust_zscore(values: np.ndarray) -> np.ndarray:
    arr = np.asarray(values, dtype=float)
    med = float(np.median(arr))
    mad = float(np.median(np.abs(arr - med)))
    if mad < 1e-12:
        return np.zeros_like(arr)
    return 0.6744897501960817 * (arr - med) / mad


def build_phi_design(phi_values: np.ndarray) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "const": 1.0,
            "phi_linear": np.asarray(phi_values, dtype=float),
            "phi_quadratic": np.asarray(phi_values, dtype=float) ** 2,
        }
    )


def fit_phi_glm(poi: np.ndarray, phi_current: np.ndarray, phi_shifted: np.ndarray):
    exog = build_phi_design(phi_current)
    glm = sm.GLM(
        poi,
        exog,
        family=sm.families.Gaussian(link=sm.families.links.Identity()),
    ).fit()
    fitted = np.asarray(glm.fittedvalues, dtype=float)
    shifted_exog = build_phi_design(phi_shifted)
    shifted_pred = np.asarray(glm.predict(shifted_exog), dtype=float)
    return glm, fitted, shifted_pred


def fit_pruned_phi_glm(
    poi: np.ndarray,
    phi_current: np.ndarray,
    phi_shifted: np.ndarray,
    *,
    outlier_z: float,
):
    del outlier_z
    glm, fitted, shifted_pred = fit_phi_glm(poi, phi_current, phi_shifted)
    resid_z = robust_zscore(poi - fitted)
    inlier_mask = np.ones(len(poi), dtype=bool)
    return glm, glm, fitted, shifted_pred, resid_z, inlier_mask


def score_batch(pair_counts: np.ndarray, combos: np.ndarray) -> np.ndarray:
    return (
        pair_counts[combos[:, 0], combos[:, 1]]
        + pair_counts[combos[:, 0], combos[:, 2]]
        + pair_counts[combos[:, 0], combos[:, 3]]
        + pair_counts[combos[:, 0], combos[:, 4]]
        + pair_counts[combos[:, 1], combos[:, 2]]
        + pair_counts[combos[:, 1], combos[:, 3]]
        + pair_counts[combos[:, 1], combos[:, 4]]
        + pair_counts[combos[:, 2], combos[:, 3]]
        + pair_counts[combos[:, 2], combos[:, 4]]
        + pair_counts[combos[:, 3], combos[:, 4]]
    ).astype(int)


def find_matching_guesses(
    pair_counts: np.ndarray,
    main_n: int,
    target_score: int,
    batch_size: int,
) -> tuple[pd.DataFrame, int]:
    combo_iter = combinations(range(1, main_n + 1), 5)
    matched_frames: list[pd.DataFrame] = []
    total = 0

    while True:
        batch = list(islice(combo_iter, batch_size))
        if not batch:
            break

        combos = np.asarray(batch, dtype=np.int16)
        scores = score_batch(pair_counts, combos)
        start_index = total + 1
        total += len(combos)

        hit_mask = scores == target_score
        if not np.any(hit_mask):
            continue

        matched = combos[hit_mask]
        matched_scores = scores[hit_mask]
        matched_positions = np.flatnonzero(hit_mask) + start_index
        matched_frames.append(
            pd.DataFrame(
                {
                    "combination_index": matched_positions.astype(np.int64),
                    "ball_1": matched[:, 0].astype(int),
                    "ball_2": matched[:, 1].astype(int),
                    "ball_3": matched[:, 2].astype(int),
                    "ball_4": matched[:, 3].astype(int),
                    "ball_5": matched[:, 4].astype(int),
                    "score": matched_scores.astype(int),
                }
            )
        )

    if matched_frames:
        return pd.concat(matched_frames, ignore_index=True), total
    return pd.DataFrame(
        columns=["combination_index", "ball_1", "ball_2", "ball_3", "ball_4", "ball_5", "score"]
    ), total


def find_matching_full7_guesses(
    pair_counts: np.ndarray,
    *,
    main_n: int,
    star_n: int,
    target_score: int,
    batch_size: int,
) -> tuple[pd.DataFrame, int]:
    main_iter = combinations(range(1, main_n + 1), 5)
    star_slots = np.arange(main_n + 1, main_n + star_n + 1, dtype=np.int16)
    star_pair_idx = np.asarray(list(combinations(range(star_n), 2)), dtype=np.int16)
    star_encoded = star_slots[star_pair_idx]
    star_display = star_pair_idx + 1
    star_scores = pair_counts[star_encoded[:, 0], star_encoded[:, 1]].astype(int)
    tickets_per_main = len(star_pair_idx)

    matched_frames: list[pd.DataFrame] = []
    processed_mains = 0

    while True:
        batch = list(islice(main_iter, batch_size))
        if not batch:
            break

        mains = np.asarray(batch, dtype=np.int16)
        main_scores = score_batch(pair_counts, mains)
        main_star_sum = pair_counts[mains[:, :, None], star_slots[None, None, :]].sum(axis=1).astype(int)

        for star_pos, (star_left_idx, star_right_idx) in enumerate(star_pair_idx):
            total_scores = (
                main_scores
                + star_scores[star_pos]
                + main_star_sum[:, star_left_idx]
                + main_star_sum[:, star_right_idx]
            )
            hit_mask = total_scores == target_score
            if not np.any(hit_mask):
                continue

            hit_positions = np.flatnonzero(hit_mask)
            combo_indices = (
                (processed_mains + hit_positions).astype(np.int64) * tickets_per_main
                + int(star_pos)
                + 1
            )
            matched_mains = mains[hit_mask]
            matched_frames.append(
                pd.DataFrame(
                    {
                        "combination_index": combo_indices,
                        "ball_1": matched_mains[:, 0].astype(int),
                        "ball_2": matched_mains[:, 1].astype(int),
                        "ball_3": matched_mains[:, 2].astype(int),
                        "ball_4": matched_mains[:, 3].astype(int),
                        "ball_5": matched_mains[:, 4].astype(int),
                        "star_1": np.full(hit_positions.shape, int(star_display[star_pos, 0]), dtype=int),
                        "star_2": np.full(hit_positions.shape, int(star_display[star_pos, 1]), dtype=int),
                        "score": total_scores[hit_mask].astype(int),
                    }
                )
            )

        processed_mains += len(mains)

    total = processed_mains * tickets_per_main
    if matched_frames:
        return pd.concat(matched_frames, ignore_index=True), total
    return pd.DataFrame(
        columns=[
            "combination_index",
            "ball_1",
            "ball_2",
            "ball_3",
            "ball_4",
            "ball_5",
            "star_1",
            "star_2",
            "score",
        ]
    ), total


def find_nearest_main5_guesses(
    pair_counts: np.ndarray,
    *,
    main_n: int,
    target_score: int,
    batch_size: int,
) -> tuple[pd.DataFrame, int]:
    combo_iter = combinations(range(1, main_n + 1), 5)
    best_gap: int | None = None
    matched_frames: list[pd.DataFrame] = []
    total = 0

    while True:
        batch = list(islice(combo_iter, batch_size))
        if not batch:
            break

        combos = np.asarray(batch, dtype=np.int16)
        scores = score_batch(pair_counts, combos)
        gaps = np.abs(scores - target_score)
        local_gap = int(gaps.min())
        start_index = total + 1
        total += len(combos)

        if best_gap is None or local_gap < best_gap:
            best_gap = local_gap
            matched_frames = []

        if local_gap != best_gap:
            continue

        hit_mask = gaps == best_gap
        matched = combos[hit_mask]
        matched_scores = scores[hit_mask]
        matched_gaps = gaps[hit_mask]
        matched_positions = np.flatnonzero(hit_mask) + start_index
        matched_frames.append(
            pd.DataFrame(
                {
                    "combination_index": matched_positions.astype(np.int64),
                    "ball_1": matched[:, 0].astype(int),
                    "ball_2": matched[:, 1].astype(int),
                    "ball_3": matched[:, 2].astype(int),
                    "ball_4": matched[:, 3].astype(int),
                    "ball_5": matched[:, 4].astype(int),
                    "score": matched_scores.astype(int),
                    "score_gap": matched_gaps.astype(int),
                }
            )
        )

    if matched_frames:
        return pd.concat(matched_frames, ignore_index=True), int(best_gap or 0)
    return pd.DataFrame(
        columns=["combination_index", "ball_1", "ball_2", "ball_3", "ball_4", "ball_5", "score", "score_gap"]
    ), -1


def find_nearest_full7_guesses(
    pair_counts: np.ndarray,
    *,
    main_n: int,
    star_n: int,
    target_score: int,
    batch_size: int,
) -> tuple[pd.DataFrame, int]:
    main_iter = combinations(range(1, main_n + 1), 5)
    star_slots = np.arange(main_n + 1, main_n + star_n + 1, dtype=np.int16)
    star_pair_idx = np.asarray(list(combinations(range(star_n), 2)), dtype=np.int16)
    star_encoded = star_slots[star_pair_idx]
    star_display = star_pair_idx + 1
    star_scores = pair_counts[star_encoded[:, 0], star_encoded[:, 1]].astype(int)
    tickets_per_main = len(star_pair_idx)

    best_gap: int | None = None
    matched_frames: list[pd.DataFrame] = []
    processed_mains = 0

    while True:
        batch = list(islice(main_iter, batch_size))
        if not batch:
            break

        mains = np.asarray(batch, dtype=np.int16)
        main_scores = score_batch(pair_counts, mains)
        main_star_sum = pair_counts[mains[:, :, None], star_slots[None, None, :]].sum(axis=1).astype(int)

        for star_pos, (star_left_idx, star_right_idx) in enumerate(star_pair_idx):
            total_scores = (
                main_scores
                + star_scores[star_pos]
                + main_star_sum[:, star_left_idx]
                + main_star_sum[:, star_right_idx]
            )
            gaps = np.abs(total_scores - target_score)
            local_gap = int(gaps.min())

            if best_gap is None or local_gap < best_gap:
                best_gap = local_gap
                matched_frames = []

            if local_gap != best_gap:
                continue

            hit_mask = gaps == best_gap
            if not np.any(hit_mask):
                continue

            hit_positions = np.flatnonzero(hit_mask)
            combo_indices = (
                (processed_mains + hit_positions).astype(np.int64) * tickets_per_main
                + int(star_pos)
                + 1
            )
            matched_mains = mains[hit_mask]
            matched_frames.append(
                pd.DataFrame(
                    {
                        "combination_index": combo_indices,
                        "ball_1": matched_mains[:, 0].astype(int),
                        "ball_2": matched_mains[:, 1].astype(int),
                        "ball_3": matched_mains[:, 2].astype(int),
                        "ball_4": matched_mains[:, 3].astype(int),
                        "ball_5": matched_mains[:, 4].astype(int),
                        "star_1": np.full(hit_positions.shape, int(star_display[star_pos, 0]), dtype=int),
                        "star_2": np.full(hit_positions.shape, int(star_display[star_pos, 1]), dtype=int),
                        "score": total_scores[hit_mask].astype(int),
                        "score_gap": gaps[hit_mask].astype(int),
                    }
                )
            )

        processed_mains += len(mains)

    if matched_frames:
        return pd.concat(matched_frames, ignore_index=True), int(best_gap or 0)
    return pd.DataFrame(
        columns=[
            "combination_index",
            "ball_1",
            "ball_2",
            "ball_3",
            "ball_4",
            "ball_5",
            "star_1",
            "star_2",
            "score",
            "score_gap",
        ]
    ), -1


def annotate_match_statistics(matches: pd.DataFrame, pair_diag: PairZDiagnostics) -> pd.DataFrame:
    if matches.empty:
        return matches.copy()

    scored = matches.copy()
    mains = scored[[f"ball_{idx}" for idx in range(1, 6)]].to_numpy(dtype=int)
    has_stars = {"star_1", "star_2"}.issubset(scored.columns)

    main_pair_sum = np.zeros(len(scored), dtype=float)
    ticket_pair_sum = np.zeros(len(scored), dtype=float)
    min_main_pair = np.full(len(scored), np.inf, dtype=float)
    min_ticket_pair = np.full(len(scored), np.inf, dtype=float)
    least_pair = pair_diag.least_main_pair
    contains_least_pair = np.zeros(len(scored), dtype=bool)

    for left_idx, right_idx in combinations(range(5), 2):
        left = mains[:, left_idx]
        right = mains[:, right_idx]
        values = pair_diag.main_main_z[left, right]
        main_pair_sum += values
        ticket_pair_sum += values
        min_main_pair = np.minimum(min_main_pair, values)
        min_ticket_pair = np.minimum(min_ticket_pair, values)
        contains_least_pair |= (
            ((left == least_pair[0]) & (right == least_pair[1]))
            | ((left == least_pair[1]) & (right == least_pair[0]))
        )

    pair_count = 10
    if has_stars:
        stars = scored[["star_1", "star_2"]].to_numpy(dtype=int)
        for main_idx in range(5):
            main_values = mains[:, main_idx]
            for star_idx in range(2):
                star_values = stars[:, star_idx]
                values = pair_diag.main_star_z[main_values, star_values]
                ticket_pair_sum += values
                min_ticket_pair = np.minimum(min_ticket_pair, values)
                pair_count += 1

        star_pair_values = pair_diag.star_star_z[stars[:, 0], stars[:, 1]]
        ticket_pair_sum += star_pair_values
        min_ticket_pair = np.minimum(min_ticket_pair, star_pair_values)
        pair_count += 1

    scored["main_pair_sum_z"] = main_pair_sum
    scored["main_pair_mean_z"] = main_pair_sum / 10.0
    scored["min_main_pair_z"] = min_main_pair
    scored["least_pair_injection_z"] = safe_zscore(min_main_pair)
    scored["contains_global_least_main_pair"] = contains_least_pair.astype(int)
    scored["ticket_pair_sum_z"] = ticket_pair_sum
    scored["ticket_pair_mean_z"] = ticket_pair_sum / float(pair_count)
    scored["min_ticket_pair_z"] = min_ticket_pair
    scored["ticket_pair_sum_zscore"] = safe_zscore(ticket_pair_sum)
    return scored


def shortlist_matches(matches: pd.DataFrame, top_n: int, *, likely: bool) -> pd.DataFrame:
    if matches.empty:
        return matches.copy()

    ranked = matches.copy()
    if likely:
        ranked = ranked.sort_values(
            ["ticket_pair_sum_zscore", "ticket_pair_mean_z", "min_ticket_pair_z"],
            ascending=[False, False, False],
        )
    else:
        ranked = ranked.sort_values(
            ["least_pair_injection_z", "min_main_pair_z", "ticket_pair_sum_zscore"],
            ascending=[True, True, True],
        )
    ranked = ranked.head(top_n).reset_index(drop=True)
    ranked.insert(0, "rank", np.arange(1, len(ranked) + 1, dtype=int))
    return ranked


def prepare_excel_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    out = df
    copied = False
    for column in df.columns:
        if np.issubdtype(df[column].dtype, np.datetime64):
            if not copied:
                out = df.copy()
                copied = True
            out[column] = out[column].dt.strftime("%Y-%m-%d")
        elif df[column].dtype == object:
            if not copied:
                out = df.copy()
                copied = True
            out[column] = out[column].map(
                lambda value: json.dumps(value)
                if isinstance(value, (list, tuple, dict))
                else value
            )
    return out


def write_excel_workbook(out_path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    max_data_rows = EXCEL_MAX_ROWS - 1

    def chunked_sheets(sheet_name: str, frame: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
        if len(frame) <= max_data_rows:
            return [(sheet_name[:31], frame)]
        chunks: list[tuple[str, pd.DataFrame]] = []
        base = sheet_name[:27]
        for chunk_idx, start in enumerate(range(0, len(frame), max_data_rows), start=1):
            stop = start + max_data_rows
            chunks.append((f"{base}_{chunk_idx}"[:31], frame.iloc[start:stop].reset_index(drop=True)))
        return chunks

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsxwriter is not None:
        workbook = xlsxwriter.Workbook(str(out_path), {"constant_memory": True})
        try:
            for sheet_name, frame in sheets:
                excel_frame = prepare_excel_frame(frame)
                for chunk_name, chunk_frame in chunked_sheets(sheet_name, excel_frame):
                    worksheet = workbook.add_worksheet(chunk_name)
                    worksheet.write_row(0, 0, list(chunk_frame.columns))
                    for row_idx, row in enumerate(chunk_frame.itertuples(index=False, name=None), start=1):
                        worksheet.write_row(row_idx, 0, row)
        finally:
            workbook.close()
        return

    if Workbook is None:
        raise RuntimeError("Neither xlsxwriter nor openpyxl is available to write the Excel workbook.")

    workbook = Workbook(write_only=True)
    for sheet_name, frame in sheets:
        excel_frame = prepare_excel_frame(frame)
        for chunk_name, chunk_frame in chunked_sheets(sheet_name, excel_frame):
            worksheet = workbook.create_sheet(title=chunk_name)
            worksheet.append(list(chunk_frame.columns))
            for row in chunk_frame.itertuples(index=False, name=None):
                worksheet.append(list(row))
    workbook.save(out_path)


def save_plot(
    frame: pd.DataFrame,
    *,
    phi_current: np.ndarray,
    fitted: np.ndarray,
    shifted_pred: np.ndarray,
    inlier_mask: np.ndarray,
    out_path: Path,
) -> None:
    fig, axes = plt.subplots(3, 1, figsize=(14, 14), constrained_layout=True)

    axes[0].plot(frame["draw_date"], frame["poi"], color="steelblue", lw=0.8, label="poi")
    axes[0].plot(frame["draw_date"], fitted, color="navy", lw=1.4, label="phi quadratic glm fitted")
    if np.any(~inlier_mask):
        axes[0].scatter(
            frame.loc[~inlier_mask, "draw_date"],
            frame.loc[~inlier_mask, "poi"],
            color="tomato",
            s=22,
            label="pruned outliers",
            zorder=3,
        )
    axes[0].set_title("Diagnostics 3: poi vs quadratic Euler-phi Gaussian GLM")
    axes[0].set_ylabel("poi")
    axes[0].legend()

    order = np.argsort(phi_current)
    axes[1].scatter(
        phi_current[inlier_mask],
        frame.loc[inlier_mask, "poi"],
        s=10,
        alpha=0.55,
        color="steelblue",
        label="inliers",
    )
    if np.any(~inlier_mask):
        axes[1].scatter(
            phi_current[~inlier_mask],
            frame.loc[~inlier_mask, "poi"],
            s=18,
            alpha=0.8,
            color="tomato",
            label="pruned outliers",
        )
    axes[1].plot(phi_current[order], fitted[order], color="tomato", lw=2, label="glm fit")
    axes[1].set_title("Euler phi vs poi (quadratic fit)")
    axes[1].set_xlabel("phi(t)")
    axes[1].set_ylabel("poi")
    axes[1].legend()

    axes[2].plot(
        frame["draw_date"].iloc[1:],
        shifted_pred[:-1],
        color="purple",
        lw=1.2,
        label="predict using phi(t+1)",
    )
    axes[2].scatter(
        frame["draw_date"].iloc[-1],
        shifted_pred[-1],
        color="tomato",
        s=45,
        label="out-of-sample phi(T+1) response",
        zorder=3,
    )
    axes[2].set_title("Shifted response predictions from g[2:length(g)]")
    axes[2].set_xlabel("draw date")
    axes[2].set_ylabel("predicted response")
    axes[2].legend()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def save_orderstat_pair_plot(
    history: pd.DataFrame,
    *,
    pair_diag: PairZDiagnostics,
    rolling_window: int,
    out_path: Path,
) -> None:
    ordered_mains = np.sort(
        history[[f"ball_{idx}" for idx in range(1, 6)]].to_numpy(dtype=float),
        axis=1,
    )
    roll = pd.DataFrame(ordered_mains, columns=[f"position_{idx}" for idx in range(1, 6)]).rolling(
        rolling_window
    ).mean()
    roll = roll.dropna().reset_index(drop=True)
    roll_dates = history["draw_date"].iloc[rolling_window - 1 :].reset_index(drop=True)

    moments = main_order_stat_moments()
    fair_means = moments.means
    fair_roll_sds = moments.stds / max(float(rolling_window) ** 0.5, 1.0)

    fig = plt.figure(figsize=(24, 16))
    gs = fig.add_gridspec(3, 5, height_ratios=[1, 1, 1.15], hspace=0.18, wspace=0.24)
    fig.suptitle("Order-Statistic & Pair Co-Occurrence Diagnostics", fontsize=20, y=0.98)

    for pos_idx in range(5):
        ax = fig.add_subplot(gs[0, pos_idx])
        values = roll.iloc[:, pos_idx].to_numpy(dtype=float)
        ax.hist(values, bins=30, density=True, alpha=0.72, color="#6f8fac", label="empirical")
        x_grid = np.linspace(values.min(), values.max(), 300)
        density = norm.pdf(x_grid, loc=fair_means[pos_idx], scale=max(fair_roll_sds[pos_idx], 1e-6))
        ax.plot(x_grid, density, color="#d04f28", lw=2, label="fair draw")
        ax.axvline(fair_means[pos_idx], color="#d04f28", ls=":", lw=1.4)
        ax.axvline(values.mean(), color="#2f5f8a", ls=":", lw=1.4)
        ax.set_title(f"position {pos_idx + 1}")
        ax.set_xlabel(f"rolling mean b({pos_idx + 1})")
        if pos_idx == 0:
            ax.set_ylabel("P[b(k) = x]")
            ax.legend(fontsize=10, frameon=False)

    for pos_idx in range(5):
        ax = fig.add_subplot(gs[1, pos_idx])
        ax.plot(roll_dates, roll.iloc[:, pos_idx], color="#4c7398", lw=1.25)
        ax.axhline(fair_means[pos_idx], color="#d04f28", ls="--", lw=1.5)
        ax.set_title(f"rolling mean b({pos_idx + 1})")
        if pos_idx == 0:
            ax.set_ylabel(f"window={rolling_window}")
        ax.tick_params(axis="x", rotation=0)

    heat = np.full((50, 50), np.nan, dtype=float)
    for left in range(1, 51):
        for right in range(1, 51):
            if left == right:
                continue
            heat[left - 1, right - 1] = pair_diag.main_main_z[left, right]

    ax_heat = fig.add_subplot(gs[2, :3])
    mesh = ax_heat.imshow(
        np.clip(heat, -3, 3),
        cmap="RdBu_r",
        vmin=-3,
        vmax=3,
        aspect="auto",
        origin="upper",
    )
    ax_heat.set_title("Pair co-occurrence z-score (clipped to ±3)\nblue = appears less than fair-draw, red = more")
    ax_heat.set_xlabel("ball j")
    ax_heat.set_ylabel("ball i")
    cbar = fig.colorbar(mesh, ax=ax_heat, fraction=0.046, pad=0.04)
    cbar.set_label("z-score")

    top15 = pair_diag.top_abs_pairs.head(15).copy()
    top15["pair_label"] = top15.apply(
        lambda row: f"({int(row['ball_1'])},{int(row['ball_2'])})",
        axis=1,
    )
    top15 = top15.sort_values("z_score", ascending=True).reset_index(drop=True)

    ax_bar = fig.add_subplot(gs[2, 3:])
    colors = np.where(top15["z_score"].to_numpy(dtype=float) >= 0.0, "#d04f28", "#4c7398")
    ax_bar.barh(top15["pair_label"], top15["z_score"], color=colors)
    ax_bar.axvline(0.0, color="black", lw=1)
    ax_bar.set_title("Top 15 |z| pairs (above expected = red, below = blue)")
    ax_bar.set_xlabel("z-score")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    args = parse_args()
    history_path = resolve_repo_path(args.history)
    out_dir = resolve_repo_path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    raw_history = load_history(history_path)
    history, effective_start_date = apply_start_date_cutoff(
        raw_history,
        mode=args.mode,
        start_date_arg=args.start_date,
    )
    pair_diag = build_pair_z_diagnostics(history)
    main_n = int(history[[f"ball_{idx}" for idx in range(1, 6)]].to_numpy(dtype=int).max())
    star_n = int(history[[f"star_{idx}" for idx in range(1, 3)]].to_numpy(dtype=int).max())
    rolling_window = max(10, min(int(args.rolling_window), len(history)))

    if args.mode == "full7":
        draws, _, _ = encode_full7_draws(history)
        pair_dimension = draws.shape[1]
        features = build_pair_features_generic(draws, universe_size=main_n + star_n, include_current=True)
    else:
        draws = history[[f"ball_{idx}" for idx in range(1, 6)]].to_numpy(dtype=int)
        pair_dimension = draws.shape[1]
        features = build_pair_features_generic(draws, universe_size=main_n, include_current=True)

    poi = features.poi.astype(float)
    phi_full = euler_phi_upto(len(poi) + 1).astype(float)
    phi_current = phi_full[:-1]
    phi_shifted = phi_full[1:]

    raw_glm, pruned_glm, fitted, shifted_pred, resid_z, inlier_mask = fit_pruned_phi_glm(
        poi,
        phi_current,
        phi_shifted,
        outlier_z=args.outlier_z,
    )
    corr = float(np.corrcoef(phi_current, poi)[0, 1])
    current_expected_poi = float(fitted[-1])
    predicted_next_poi = float(shifted_pred[-1])
    predicted_growth_n1 = predicted_next_poi - current_expected_poi
    reverse_growth_target_value = (2.0 * current_expected_poi) - predicted_next_poi
    reverse_growth_target_score = int(round(reverse_growth_target_value))

    r_window = min(6, len(poi))
    trailing_mean_r = float(np.mean(poi[-r_window:]))
    trailing_mean_last_5 = float(np.mean(poi[-min(5, len(poi)) :]))
    target_score = int(round(trailing_mean_r))

    series = pd.DataFrame(
        {
            "draw_date": history["draw_date"],
            "poi": poi,
            "phi": phi_current,
            "phi_resid_robust_z": resid_z,
            "phi_inlier": inlier_mask.astype(int),
            "glm_fitted_pruned": fitted,
            "glm_resid": poi - fitted,
            "phi_next": phi_shifted,
            "glm_predicted_next_response": shifted_pred,
        }
    )
    series.to_csv(out_dir / "diagnostics3_series.csv", index=False)

    if args.mode == "full7":
        matches, total_scored = find_matching_full7_guesses(
            features.pair_counts,
            main_n=main_n,
            star_n=star_n,
            target_score=target_score,
            batch_size=args.batch_size,
        )
        reverse_matches, _ = find_matching_full7_guesses(
            features.pair_counts,
            main_n=main_n,
            star_n=star_n,
            target_score=reverse_growth_target_score,
            batch_size=args.batch_size,
        )
        if reverse_matches.empty:
            reverse_matches, reverse_gap = find_nearest_full7_guesses(
                features.pair_counts,
                main_n=main_n,
                star_n=star_n,
                target_score=reverse_growth_target_score,
                batch_size=args.batch_size,
            )
        else:
            reverse_gap = 0
    else:
        matches, total_scored = find_matching_guesses(
            features.pair_counts,
            main_n=main_n,
            target_score=target_score,
            batch_size=args.batch_size,
        )
        reverse_matches, _ = find_matching_guesses(
            features.pair_counts,
            main_n=main_n,
            target_score=reverse_growth_target_score,
            batch_size=args.batch_size,
        )
        if reverse_matches.empty:
            reverse_matches, reverse_gap = find_nearest_main5_guesses(
                features.pair_counts,
                main_n=main_n,
                target_score=reverse_growth_target_score,
                batch_size=args.batch_size,
            )
        else:
            reverse_gap = 0

    matches = annotate_match_statistics(matches, pair_diag)
    reverse_matches = annotate_match_statistics(reverse_matches, pair_diag)

    matches.to_csv(out_dir / "diagnostics3_guesses.csv", index=False)
    reverse_matches.to_csv(out_dir / "diagnostics3_reverse_growth_matches.csv", index=False)
    pair_diag.top_abs_pairs.to_csv(out_dir / "diagnostics3_main_pair_zscores.csv", index=False)

    least_pair_shortlist = shortlist_matches(matches, args.excel_top_n, likely=False)
    super_likely_shortlist = shortlist_matches(reverse_matches, args.excel_top_n, likely=True)
    reverse_effective_score = (
        int(super_likely_shortlist["score"].iloc[0])
        if not super_likely_shortlist.empty
        else reverse_growth_target_score
    )
    super_likely_shortlist["selection_source"] = (
        "reverse_growth_exact" if reverse_gap == 0 else "reverse_growth_nearest"
    )
    super_likely_shortlist["target_score_used"] = reverse_effective_score
    super_likely_shortlist["reverse_target_gap"] = reverse_gap

    least_pair_shortlist["selection_source"] = "least_pair_injection"
    least_pair_shortlist["target_score_used"] = target_score

    save_plot(
        series,
        phi_current=phi_current,
        fitted=fitted,
        shifted_pred=shifted_pred,
        inlier_mask=inlier_mask,
        out_path=out_dir / "diagnostics3_glm.png",
    )
    save_orderstat_pair_plot(
        history,
        pair_diag=pair_diag,
        rolling_window=rolling_window,
        out_path=out_dir / "diagnostics3_orderstat_pairs.png",
    )

    summary = Diagnostics3Summary(
        raw_rows=len(raw_history),
        mode=args.mode,
        rows=len(series),
        cutoff_start_date=effective_start_date,
        start_date=history["draw_date"].min().date().isoformat(),
        end_date=history["draw_date"].max().date().isoformat(),
        main_n=main_n,
        star_n=star_n,
        pair_dimension=pair_dimension,
        phi_poi_corr=corr,
        poi_mean=float(poi.mean()),
        poi_std=float(poi.std(ddof=0)),
        phi_model="quadratic",
        glm_family="gaussian",
        glm_link="identity",
        model_intercept=float(pruned_glm.params["const"]),
        model_phi_linear=float(pruned_glm.params["phi_linear"]),
        model_phi_quadratic=float(pruned_glm.params["phi_quadratic"]),
        model_aic=float(pruned_glm.aic),
        model_deviance=float(pruned_glm.deviance),
        model_fitted_mean=float(fitted.mean()),
        model_shifted_prediction_mean=float(shifted_pred.mean()),
        pruning_applied=False,
        pruning_z_threshold=float(args.outlier_z),
        inlier_count=int(inlier_mask.sum()),
        outlier_count=int((~inlier_mask).sum()),
        current_expected_poi=current_expected_poi,
        predicted_next_poi=predicted_next_poi,
        predicted_growth_n1=predicted_growth_n1,
        reverse_growth_target_score=reverse_growth_target_score,
        reverse_growth_target_value=float(reverse_growth_target_value),
        reverse_growth_effective_score=int(reverse_effective_score),
        reverse_growth_score_gap=int(reverse_gap),
        trailing_window_r_compatible=r_window,
        trailing_mean_r_compatible=trailing_mean_r,
        trailing_mean_last_5=trailing_mean_last_5,
        target_score_r_compatible=target_score,
        matching_guess_count=int(len(matches)),
        total_combinations_scored=total_scored,
        reverse_growth_match_count=int(len(reverse_matches)),
        least_likely_main_pair=[int(pair_diag.least_main_pair[0]), int(pair_diag.least_main_pair[1])],
        least_likely_main_pair_z=float(pair_diag.least_main_pair_z),
        most_likely_main_pair=[int(pair_diag.most_main_pair[0]), int(pair_diag.most_main_pair[1])],
        most_likely_main_pair_z=float(pair_diag.most_main_pair_z),
    )
    (out_dir / "diagnostics3_summary.json").write_text(
        json.dumps(asdict(summary), indent=2),
        encoding="utf-8",
    )

    summary_frame = pd.DataFrame([asdict(summary)])
    excel_path = out_dir / "diagnostics3_candidates.xlsx"
    write_excel_workbook(
        excel_path,
        [
            ("summary", summary_frame),
            ("growth_matches", matches),
            ("super_likely", super_likely_shortlist),
            ("least_injection", least_pair_shortlist),
            ("pair_zscores", pair_diag.top_abs_pairs),
        ],
    )

    print("=" * 70)
    print("DIAGNOSTICS 3 — Euler Phi Gaussian GLM")
    print("=" * 70)
    print(
        f"Rows: {summary.rows} of {summary.raw_rows}  "
        f"Cutoff: {summary.cutoff_start_date}  "
        f"Date range: {summary.start_date} -> {summary.end_date}  "
        f"mode={summary.mode}  pair_dim={summary.pair_dimension}"
    )
    print(f"Correlation(phi, poi): {summary.phi_poi_corr:.4f}")
    print(
        f"GLM: poi ~ phi(t) + phi(t)^2  family={summary.glm_family}  link={summary.glm_link}  "
        f"b0={summary.model_intercept:.4f}  "
        f"b1={summary.model_phi_linear:.4f}  "
        f"b2={summary.model_phi_quadratic:.6f}"
    )
    print(
        f"Residual pruning applied={summary.pruning_applied}  "
        f"inliers={summary.inlier_count}  outliers={summary.outlier_count}"
    )
    print(
        f"R-compatible trailing mean window={summary.trailing_window_r_compatible} draws  "
        f"mean={summary.trailing_mean_r_compatible:.3f}  "
        f"target score={summary.target_score_r_compatible}"
    )
    print(
        f"Reverse-growth target: value={summary.reverse_growth_target_value:.3f}  "
        f"score={summary.reverse_growth_target_score}  "
        f"effective score={summary.reverse_growth_effective_score}  "
        f"gap={summary.reverse_growth_score_gap}  "
        f"matches={summary.reverse_growth_match_count}"
    )
    print(
        f"Least likely main pair: {tuple(summary.least_likely_main_pair)}  "
        f"z={summary.least_likely_main_pair_z:.3f}"
    )
    print(f"Matching guesses: {summary.matching_guess_count} out of {summary.total_combinations_scored}")
    print(f"Wrote: {out_dir / 'diagnostics3_series.csv'}")
    print(f"Wrote: {out_dir / 'diagnostics3_guesses.csv'}")
    print(f"Wrote: {out_dir / 'diagnostics3_reverse_growth_matches.csv'}")
    print(f"Wrote: {out_dir / 'diagnostics3_main_pair_zscores.csv'}")
    print(f"Wrote: {out_dir / 'diagnostics3_summary.json'}")
    print(f"Wrote: {out_dir / 'diagnostics3_glm.png'}")
    print(f"Wrote: {out_dir / 'diagnostics3_orderstat_pairs.png'}")
    print(f"Wrote: {excel_path}")


if __name__ == "__main__":
    main()
